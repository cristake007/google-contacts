#!/usr/bin/env python3
"""Find the most likely company domain with one Brave Search API query per row."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


API_URL = "https://api.search.brave.com/res/v1/web/search"
DEFAULT_INPUT = Path(__file__).with_name("all_cleaned.xlsx")
DEFAULT_OUTPUT = Path(__file__).with_name("all_cleaned_brave_domains.xlsx")

COMPANY_HEADER = "denumire_companie"
ADDRESS_HEADER = "adresa_punctului_de_lucru"
CUI_HEADERS = ("cui_clean", "cod_unic_inregistrare")

OUTPUT_HEADERS = (
    "brave_domain",
    "brave_url",
    "brave_confidence",
    "brave_score",
    "brave_result_rank",
    "brave_result_title",
    "brave_query",
    "brave_status",
    "brave_checked_at",
    "brave_notes",
)

# These sources can mention a company, but their own domain is not the company's
# official web domain.
EXCLUDED_DOMAINS = {
    "2gis.com",
    "anaf.ro",
    "apple.com",
    "bizoo.ro",
    "cylex.ro",
    "facebook.com",
    "firme.info",
    "firmadeaur.ro",
    "google.com",
    "instagram.com",
    "linkedin.com",
    "listafirme.ro",
    "mfinante.gov.ro",
    "paginiaurii.ro",
    "risco.ro",
    "romanian-companies.eu",
    "termene.ro",
    "tiktok.com",
    "tripadvisor.com",
    "wikipedia.org",
    "x.com",
    "youtube.com",
}

LEGAL_TOKENS = {
    "co",
    "compania",
    "company",
    "grup",
    "group",
    "intreprindere",
    "pfa",
    "ra",
    "sa",
    "sc",
    "srl",
    "srl-d",
}

ADDRESS_NOISE_TOKENS = {
    "adresa",
    "ap",
    "apartament",
    "bl",
    "bloc",
    "calea",
    "com",
    "comuna",
    "et",
    "etaj",
    "jud",
    "judet",
    "loc",
    "localitate",
    "nr",
    "numar",
    "romania",
    "sat",
    "sc",
    "str",
    "strada",
    "sector",
}


@dataclass(frozen=True)
class Candidate:
    domain: str
    url: str
    title: str
    rank: int
    score: int
    occurrences: int = 1


@dataclass(frozen=True)
class SearchJob:
    row: int
    company: str
    address: str
    query: str


@dataclass(frozen=True)
class SearchOutcome:
    job: SearchJob
    values: dict[str, Any]
    stop_run: bool = False


class BraveApiError(RuntimeError):
    def __init__(self, message: str, *, stop_run: bool = False) -> None:
        super().__init__(message)
        self.stop_run = stop_run


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(re.findall(r"[a-z0-9]+", text.casefold()))


def tokens(value: Any) -> list[str]:
    return normalize_text(value).split()


def company_tokens(company: str) -> list[str]:
    return [
        token
        for token in tokens(company)
        if token not in LEGAL_TOKENS and (len(token) >= 3 or token.isdigit())
    ]


def address_tokens(address: str) -> list[str]:
    return [
        token
        for token in tokens(address)
        if token not in ADDRESS_NOISE_TOKENS
        and len(token) >= 3
        and not token.isdigit()
    ]


def canonical_domain(url: str) -> str:
    parsed = urlparse(url if "://" in url else f"https://{url}")
    domain = (parsed.hostname or "").casefold().strip(".")
    if domain.startswith("www."):
        domain = domain[4:]
    try:
        domain = domain.encode("ascii").decode("idna")
    except (UnicodeError, UnicodeDecodeError):
        pass
    return domain


def is_excluded_domain(domain: str) -> bool:
    return any(domain == item or domain.endswith(f".{item}") for item in EXCLUDED_DOMAINS)


def coverage(expected: list[str], actual_text: str) -> float:
    if not expected:
        return 0.0
    actual = set(tokens(actual_text))
    return sum(token in actual for token in set(expected)) / len(set(expected))


def score_result(
    company: str,
    address: str,
    result: dict[str, Any],
    rank: int,
) -> Candidate | None:
    url = str(result.get("url") or "").strip()
    domain = canonical_domain(url)
    if not domain or is_excluded_domain(domain):
        return None

    title = str(result.get("title") or "").strip()
    description = str(result.get("description") or "").strip()
    context = f"{title} {description} {url}"
    expected_company = company_tokens(company)
    expected_address = address_tokens(address)
    domain_words = normalize_text(domain.replace(".", " ").replace("-", " "))
    domain_compact = re.sub(r"[^a-z0-9]", "", domain.split(".", 1)[0])
    company_compact = "".join(expected_company)

    score = max(0, 30 - (rank - 1) * 2)
    score += round(42 * coverage(expected_company, context))
    score += round(35 * coverage(expected_company, domain_words))
    score += round(18 * coverage(expected_address[:8], context))

    if company_compact and len(company_compact) >= 4:
        if company_compact == domain_compact:
            score += 38
        elif company_compact in domain_compact or domain_compact in company_compact:
            score += 24

    street_numbers = re.findall(r"\b\d+[a-z]?\b", normalize_text(address))
    if street_numbers and any(number in tokens(context) for number in street_numbers):
        score += 8
    if domain.endswith(".ro"):
        score += 6

    return Candidate(
        domain=domain,
        url=url,
        title=title,
        rank=rank,
        score=score,
    )


def select_candidate(
    company: str,
    address: str,
    results: list[dict[str, Any]],
) -> tuple[Candidate | None, list[Candidate]]:
    best_by_domain: dict[str, Candidate] = {}
    counts: dict[str, int] = {}

    for rank, result in enumerate(results, start=1):
        candidate = score_result(company, address, result, rank)
        if candidate is None:
            continue
        counts[candidate.domain] = counts.get(candidate.domain, 0) + 1
        previous = best_by_domain.get(candidate.domain)
        if previous is None or candidate.score > previous.score:
            best_by_domain[candidate.domain] = candidate

    candidates = [
        Candidate(
            domain=item.domain,
            url=item.url,
            title=item.title,
            rank=item.rank,
            score=item.score + min(12, (counts[item.domain] - 1) * 4),
            occurrences=counts[item.domain],
        )
        for item in best_by_domain.values()
    ]
    candidates.sort(key=lambda item: (-item.score, item.rank, item.domain))
    return (candidates[0] if candidates else None), candidates


def confidence_for(candidate: Candidate, runner_up: Candidate | None) -> str:
    margin = candidate.score - runner_up.score if runner_up else candidate.score
    if candidate.score >= 105 and margin >= 12:
        return "HIGH"
    if candidate.score >= 75 and margin >= 5:
        return "MEDIUM"
    return "LOW"


def brave_search(api_key: str, query: str, timeout: float) -> list[dict[str, Any]]:
    # This function intentionally performs exactly one request and has no retry
    # loop, so a company can never consume more than one Brave query per run.
    parameters = urlencode(
        {
            "q": query,
            # Brave does not offer RO as a web-search market. ALL avoids the
            # default US market, while X-Loc-Country below supplies the
            # Romanian location bias using the full ISO country list.
            "country": "ALL",
            "search_lang": "ro",
            "count": 20,
            "result_filter": "web",
            "text_decorations": "false",
            "safesearch": "moderate",
        }
    )
    request = Request(
        f"{API_URL}?{parameters}",
        headers={
            "Accept": "application/json",
            "Accept-Encoding": "identity",
            "X-Subscription-Token": api_key,
            "X-Loc-Country": "RO",
            "User-Agent": "brave-domain-finder/1.0",
        },
    )

    try:
        with urlopen(request, timeout=timeout) as response:
            payload = json.load(response)
    except HTTPError as error:
        response_body = error.read(2_048).decode("utf-8", errors="replace").strip()
        detail = ""
        if response_body:
            try:
                error_payload = json.loads(response_body)
                api_error = error_payload.get("error", {})
                if isinstance(api_error, dict):
                    detail = str(api_error.get("detail") or api_error.get("message") or "").strip()
                    if not detail and api_error.get("meta"):
                        detail = json.dumps(api_error["meta"], ensure_ascii=False, separators=(",", ":"))
                if not detail:
                    detail = str(error_payload.get("message") or "").strip()
            except (json.JSONDecodeError, AttributeError):
                detail = response_body
        detail = re.sub(r"\s+", " ", detail)[:500]
        message = f"Brave API returned HTTP {error.code}"
        if detail:
            message += f": {detail}"
        stop_run = 400 <= error.code < 500
        raise BraveApiError(message, stop_run=stop_run) from error
    except (URLError, TimeoutError, json.JSONDecodeError) as error:
        raise BraveApiError(f"Brave API request failed: {error}") from error

    results = payload.get("web", {}).get("results", [])
    if not isinstance(results, list):
        raise BraveApiError("Brave API response did not contain a valid web results list")
    return [item for item in results if isinstance(item, dict)]


def header_map(worksheet) -> dict[str, int]:
    return {
        normalize_text(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }


def require_column(headers: dict[str, int], name: str) -> int:
    column = headers.get(normalize_text(name))
    if column is None:
        raise ValueError(f"Required workbook column not found: {name}")
    return column


def ensure_output_columns(worksheet) -> dict[str, int]:
    headers = header_map(worksheet)
    columns: dict[str, int] = {}
    for name in OUTPUT_HEADERS:
        normalized = normalize_text(name)
        column = headers.get(normalized)
        if column is None:
            column = worksheet.max_column + 1
            cell = worksheet.cell(1, column, name)
            cell.font = Font(bold=True)
            headers[normalized] = column
        columns[name] = column
    return columns


def cell_text(worksheet, row: int, column: int | None) -> str:
    if column is None:
        return ""
    return str(worksheet.cell(row, column).value or "").strip()


def save_workbook(workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.stem}.tmp{output_path.suffix}")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    except PermissionError as error:
        raise PermissionError(
            f"Cannot safely replace {output_path}. Close the workbook in Excel and run again."
        ) from error
    except KeyboardInterrupt:
        # Finish a clean atomic save before honoring Ctrl+C. The existing output
        # remains intact until os.replace succeeds.
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
        raise


def write_values(worksheet, row: int, columns: dict[str, int], values: dict[str, Any]) -> None:
    for name in OUTPUT_HEADERS:
        worksheet.cell(row, columns[name], values.get(name, ""))


def load_env_file(path: Path) -> None:
    """Load simple KEY=VALUE entries without replacing existing environment values."""
    if not path.exists():
        return
    try:
        lines = path.read_text(encoding="utf-8-sig").splitlines()
    except OSError as error:
        raise ValueError(f"Could not read environment file {path}: {error}") from error

    for line_number, raw_line in enumerate(lines, start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[7:].lstrip()
        if "=" not in line:
            raise ValueError(f"Invalid .env entry on line {line_number}: expected KEY=VALUE")
        name, value = line.split("=", 1)
        name = name.strip()
        if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name):
            raise ValueError(f"Invalid .env variable name on line {line_number}")
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {'"', "'"}:
            value = value[1:-1]
        os.environ.setdefault(name, value)


def search_company(api_key: str, timeout: float, job: SearchJob) -> SearchOutcome:
    checked_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    try:
        # This is the only brave_search call for this job. A SearchJob is
        # submitted once, and brave_search itself never retries.
        results = brave_search(api_key, job.query, timeout)
        selected, candidates = select_candidate(job.company, job.address, results)
        if selected is None:
            values = {
                "brave_query": job.query,
                "brave_status": "NO_DOMAIN",
                "brave_checked_at": checked_at,
                "brave_notes": f"No eligible company domain in {len(results)} web results",
            }
        else:
            runner_up = candidates[1] if len(candidates) > 1 else None
            confidence = confidence_for(selected, runner_up)
            alternatives = ", ".join(
                f"{item.domain} ({item.score})" for item in candidates[1:4]
            )
            values = {
                "brave_domain": selected.domain,
                "brave_url": selected.url,
                "brave_confidence": confidence,
                "brave_score": selected.score,
                "brave_result_rank": selected.rank,
                "brave_result_title": selected.title,
                "brave_query": job.query,
                "brave_status": "FOUND" if confidence != "LOW" else "REVIEW_LOW_CONFIDENCE",
                "brave_checked_at": checked_at,
                "brave_notes": f"Alternatives: {alternatives}" if alternatives else "",
            }
        return SearchOutcome(job=job, values=values)
    except BraveApiError as error:
        return SearchOutcome(
            job=job,
            values={
                "brave_query": job.query,
                "brave_status": "API_ERROR",
                "brave_checked_at": checked_at,
                "brave_notes": str(error),
            },
            stop_run=error.stop_run,
        )


def run_batch(
    api_key: str,
    timeout: float,
    jobs: list[SearchJob],
) -> tuple[list[SearchOutcome], bool]:
    """Run each job once and return completed outcomes plus interrupt state."""
    executor = ThreadPoolExecutor(max_workers=len(jobs), thread_name_prefix="brave-search")
    future_jobs: dict[Future[SearchOutcome], SearchJob] = {
        executor.submit(search_company, api_key, timeout, job): job for job in jobs
    }
    outcomes: list[SearchOutcome] = []
    collected: set[Future[SearchOutcome]] = set()
    interrupted = False

    try:
        for future in as_completed(future_jobs):
            outcomes.append(future.result())
            collected.add(future)
    except KeyboardInterrupt:
        interrupted = True
        for future in future_jobs:
            future.cancel()
    finally:
        # Running HTTP calls cannot be killed safely. Wait at most for their
        # configured HTTP timeouts, then retain every completed result.
        executor.shutdown(wait=True, cancel_futures=True)

    if interrupted:
        for future in future_jobs:
            if future not in collected and future.done() and not future.cancelled():
                outcomes.append(future.result())

    outcomes.sort(key=lambda outcome: outcome.job.row)
    return outcomes, interrupted


def execute_and_save_batch(
    api_key: str,
    timeout: float,
    jobs: list[SearchJob],
    worksheet,
    output_columns: dict[str, int],
    workbook,
    output_path: Path,
) -> tuple[int, bool, bool]:
    print(f"Starting batch of {len(jobs)} companies")
    for job in jobs:
        print(f"  [Excel row {job.row}] {job.query}")

    outcomes, interrupted = run_batch(api_key, timeout, jobs)
    stop_run = False
    for outcome in outcomes:
        write_values(worksheet, outcome.job.row, output_columns, outcome.values)
        status = str(outcome.values.get("brave_status") or "")
        domain = str(outcome.values.get("brave_domain") or "-")
        print(f"  row {outcome.job.row}: {status} | domain={domain}")
        if status == "API_ERROR":
            print(f"    {outcome.values.get('brave_notes', '')}", file=sys.stderr)
        stop_run = stop_run or outcome.stop_run

    # Only the main thread touches openpyxl. A batch becomes durable before the
    # next batch is allowed to start.
    save_workbook(workbook, output_path)
    print(f"Batch saved: {len(outcomes)}/{len(jobs)} completed")
    return len(outcomes), interrupted, stop_run


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read all_cleaned.xlsx and choose the most likely company domain "
            "from exactly one Brave Search API query per company."
        )
    )
    parser.add_argument("input", nargs="?", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--resume", action="store_true", help="Continue from an existing output workbook")
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional lower query limit; 0 means use --batch-size",
    )
    parser.add_argument("--only-cui", help="Process only the row with this CUI")
    parser.add_argument(
        "--batch-size",
        type=int,
        default=5,
        help="Maximum companies searched in this run (default: 5)",
    )
    parser.add_argument("--timeout", type=float, default=30.0, help="HTTP timeout in seconds")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if (
        args.limit < 0
        or not 1 <= args.batch_size <= 50
        or args.timeout <= 0
    ):
        print(
            "--limit cannot be negative; --batch-size must be 1-50; "
            "--timeout must be positive",
            file=sys.stderr,
        )
        return 2

    try:
        load_env_file(Path(__file__).with_name(".env"))
    except ValueError as error:
        print(str(error), file=sys.stderr)
        return 2

    api_key = os.environ.get("BRAVE_API_KEY", "").strip()
    if not api_key:
        print("Set BRAVE_API_KEY in .env or in the environment before running the script.", file=sys.stderr)
        return 2

    input_path = args.input.resolve()
    output_path = args.output.resolve()
    workbook_path = output_path if args.resume and output_path.exists() else input_path
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    if input_path == output_path and not args.resume:
        print("The output path must differ from the input path.", file=sys.stderr)
        return 2

    try:
        workbook = load_workbook(workbook_path)
        worksheet = workbook.active
        headers = header_map(worksheet)
        company_column = require_column(headers, COMPANY_HEADER)
        address_column = require_column(headers, ADDRESS_HEADER)
        cui_column = require_column(headers, CUI_HEADERS[0])
        output_columns = ensure_output_columns(worksheet)
    except (OSError, ValueError) as error:
        print(str(error), file=sys.stderr)
        return 2

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    try:
        # Fail before consuming quota if Excel has the output file locked.
        save_workbook(workbook, output_path)
    except PermissionError as error:
        print(str(error), file=sys.stderr)
        return 2
    except KeyboardInterrupt:
        print("\nStopped before any queries were made.", file=sys.stderr)
        return 130

    queries_scheduled = 0
    queries_completed = 0
    stopped_early = False
    interrupted = False
    pending_jobs: list[SearchJob] = []
    run_limit = min(args.limit, args.batch_size) if args.limit else args.batch_size

    try:
        for row in range(2, worksheet.max_row + 1):
            if queries_scheduled >= run_limit:
                break
            if args.only_cui and cell_text(worksheet, row, cui_column) != args.only_cui.strip():
                continue
            if args.resume and cell_text(worksheet, row, output_columns["brave_status"]):
                continue

            company = cell_text(worksheet, row, company_column)
            address = cell_text(worksheet, row, address_column)
            cui = cell_text(worksheet, row, cui_column)
            if not company or not address:
                write_values(
                    worksheet,
                    row,
                    output_columns,
                    {
                        "brave_status": "SKIPPED_MISSING_INPUT",
                        "brave_notes": "Both company name and address are required",
                    },
                )
                continue

            query = " ".join(value for value in (company, address, cui) if value)
            pending_jobs.append(
                SearchJob(
                    row=row,
                    company=company,
                    address=address,
                    query=query,
                )
            )
            queries_scheduled += 1

        if pending_jobs and not interrupted and not stopped_early:
            completed, interrupted, stopped_early = execute_and_save_batch(
                api_key,
                args.timeout,
                pending_jobs,
                worksheet,
                output_columns,
                workbook,
                output_path,
            )
            queries_completed += completed
    except KeyboardInterrupt:
        save_workbook(workbook, output_path)
        interrupted = True
    except PermissionError as error:
        print(str(error), file=sys.stderr)
        return 2

    try:
        save_workbook(workbook, output_path)
    except PermissionError as error:
        print(str(error), file=sys.stderr)
        return 2
    except KeyboardInterrupt:
        interrupted = True
    if interrupted:
        print(f"\nStopped safely; {queries_completed} completed queries were saved to {output_path}", file=sys.stderr)
        return 130
    if stopped_early:
        print("Stopping after an authentication, validation, or rate-limit error.", file=sys.stderr)
    print(f"Saved {queries_completed} queried companies to {output_path}")
    return 1 if stopped_early else 0


if __name__ == "__main__":
    raise SystemExit(main())

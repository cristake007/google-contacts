#!/usr/bin/env python3
"""
Google-only company website and contact finder for Windows.

Focused workflow:

1. Search Google using company name plus location.
2. Check Google's right-side company/knowledge panel FIRST for its Website link.
3. If no reliable panel website exists, inspect the organic Google results.
4. Select one likely official website or an exact panel-linked business listing.
5. Open only that landing page and, for official sites, one contact page.
6. Extract contacts from the listing, contact page, or homepage footer.
7. Save progress after every company and support resume.

Accuracy safeguards:

- Company-information sites, social networks and government portals are always
  excluded. Only allowlisted contact directories may pass full identity checks.
- Ambiguous names such as APICOLA or MATCA require location or CUI evidence.
- A domain already assigned to another CUI is not silently reused.
- Google result rank alone can never prove that a website is official.
- The script never crawls legal, GDPR, privacy, terms or about pages.
- CAPTCHA is resolved manually in the visible Chromium window.

Expected Excel columns:
- denumire_companie (or supported alias)
- cui_clean (or supported alias)

Optional columns used for disambiguation:
- judet
- adresa_punctului_de_lucru

PowerShell examples:

    & ".\\.venv\\Scripts\\python.exe" .\\google_contact_finder_v4.py `
      .\\all_cleaned.xlsx --only-cui 256

    & ".\\.venv\\Scripts\\python.exe" .\\google_contact_finder_v4.py `
      .\\all_cleaned.xlsx --limit 20

    & ".\\.venv\\Scripts\\python.exe" .\\google_contact_finder_v4.py `
      .\\all_cleaned.xlsx --resume --limit 20

First-time setup:
    .\bootstrap.ps1

See README.md for fresh-run, resume, and retry commands.
"""

from __future__ import annotations

import argparse
import re
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, urljoin, urlparse, urlunparse

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import (
    BrowserContext,
    Locator,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

VERSION = "4.7"
PROFILE_DIR = Path(__file__).parent / ".browser-profile"

COMPANY_ALIASES = (
    "denumire_companie",
    "denumire",
    "nume_firma",
    "firma",
    "company",
)
CUI_ALIASES = (
    "cui_clean",
    "cui",
    "cod_unic_inregistrare",
    "cod_fiscal",
)
COUNTY_ALIASES = ("judet", "județ", "county")
ADDRESS_ALIASES = (
    "adresa_punctului_de_lucru",
    "adresa",
    "adresă",
    "address",
)

OUTPUT_COLUMNS = (
    "google_website",
    "google_source",
    "google_result_rank",
    "google_result_title",
    "google_query_used",
    "google_panel_context",
    "website_cui",
    "cui_match_status",
    "contact_page_url",
    "contact_email",
    "contact_phone",
    "contact_all_emails",
    "contact_all_phones",
    "contact_found_in",
    "contact_status",
    "website_score",
    "duplicate_domain_cui",
    "contact_checked_at",
    "contact_notes",
)

STATUS_FILL_COLORS = {
    "FOUND_CONTACT_PAGE": "D9EAD3",
    "FOUND_FOOTER": "D9EAD3",
    "FOUND_PANEL_LISTING": "D9EAD3",
    "FOUND_VERIFIED_LISTING": "D9EAD3",
    "FOUND_GOOGLE_PANEL": "D9EAD3",
    "WEBSITE_NO_CONTACT": "D9EAF7",
    "REVIEW_GOOGLE_CANDIDATE": "FCE5CD",
    "REVIEW_DUPLICATE_DOMAIN": "FCE5CD",
    "REVIEW_AMBIGUOUS_NAME": "FCE5CD",
    "REVIEW_CUI_MISMATCH": "FCE5CD",
    "REVIEW_PANEL_LISTING": "FCE5CD",
    "REVIEW_IDENTITY_MISMATCH": "FCE5CD",
    "NO_WEBSITE": "E7E6E6",
    "GOOGLE_BLOCKED": "FFF2CC",
    "ERROR": "F4CCCC",
}

# Sources that may mention a company but are not its official website.
EXCLUDED_DOMAINS = {
    # Google and social platforms.
    "google.com",
    "google.ro",
    "googleusercontent.com",
    "youtube.com",
    "facebook.com",
    "instagram.com",
    "linkedin.com",
    "tiktok.com",
    "x.com",
    "twitter.com",

    # Romanian/international company directories and listing sites.
    "paginiaurii.ro",
    "listafirme.ro",
    "listafirme.eu",
    "firme.ro",
    "firmeo.ro",
    "firme-on-line.ro",
    "firme.info",
    "firmero.ro",
    "finfo.ro",
    "cautarefirme.ro",
    "termene.ro",
    "risco.ro",
    "confidas.ro",
    "targetare.ro",
    "topfirme.com",
    "totalfirme.ro",
    "infoquick.ro",
    "bonitate.ro",
    "cylex.ro",
    "firmania.ro",
    "romanian-companies.eu",
    "europages.ro",
    "kompass.com",
    "lege5.ro",
    "pet-box.ro",
    "pharmacompass.com",
    "emis.com",
    "cauta.vet",
    "cabinetvet.ro",
    "canina.ro",
    "ghidulvet.ro",
    "ghidulveterinarilor.ro",
    "medatlas.ro",
    "catalogafaceri.ro",
    "cataloage.ro",
    "consultanti.ro",
    "demoanaf.ro",
    "deschis.ro",
    "firmeapi.ro",
    "maptons.com",
    "olx.ro",
    "quickconta.ro",
    "rolocal.ro",
    "metricbiz.ro",
    "datasrl.ro",
    "firmoteca.ro",
    "icapb2b.ro",
    "constatator-urgent.ro",
    "tendersight.ai",
    "rrf.ro",
    "wikimapia.org",
    "waze.com",

    # Government, professional and procurement portals.
    "anaf.ro",
    "onrc.ro",
    "data.gov.ro",
    "gov.ro",
    "just.ro",
    "mfinante.gov.ro",
    "ansvsa.ro",
    "dsvsa.ro",
    "sicap.ai",
    "e-licitatie.ro",
    "cmvro.ro",
    "aca.org.ro",

    # Generic content/certification/repository sites observed in tests.
    "snapcert.ro",
    "digidemat.ro",
    "clientsolutions.io",
    "brat.ro",
    "leafletjs.com",
    "fcrmedia.ro",
    "rentasite.ro",
    "bionestcluster.ro",
}

# Unlike financial/company-data aggregators, these local directories primarily
# expose public contact details. They may be used only after identity checks.
VERIFIED_CONTACT_DIRECTORIES = {
    "paginiaurii.ro",
    "cylex.ro",
    "deschis.ro",
}

LEGAL_SUFFIXES = {
    "srl",
    "sa",
    "sc",
    "srl-d",
    "pfa",
    "ii",
    "if",
    "snc",
    "sca",
    "sapa",
    "ra",
    "romania",
    "românia",
    "company",
    "compania",
    "societatea",
    "comerciala",
    "comercială",
    "group",
    "grup",
    "holding",
}

# Tokens too broad to prove that a domain belongs to a company.
GENERIC_DOMAIN_WORDS = {
    "company",
    "group",
    "holding",
    "romania",
    "medical",
    "medic",
    "pharma",
    "farm",
    "vet",
    "veterinar",
    "construct",
    "consult",
    "service",
    "services",
    "trade",
    "import",
    "export",
    "apicola",
    "agricola",
    "matca",
}

GENERIC_EMAIL_PREFIXES = (
    "office",
    "contact",
    "info",
    "sales",
    "vanzari",
    "vânzări",
    "comercial",
    "secretariat",
    "receptie",
    "recepție",
    "administrativ",
    "support",
    "hello",
    "comenzi",
    "clienti",
    "clienți",
)

LOW_PRIORITY_EMAIL_PREFIXES = (
    "dpo",
    "gdpr",
    "privacy",
    "confidentialitate",
    "confidențialitate",
    "noreply",
    "no-reply",
    "donotreply",
    "webmaster",
    "abuse",
    "jobs",
    "job",
    "career",
    "cariere",
)

TECHNICAL_EMAIL_DOMAINS = {
    "example.com",
    "example.org",
    "sentry.io",
    "wixpress.com",
    "wordpress.org",
    "cloudflare.com",
    "schema.org",
    "w3.org",
}

CONTACT_LINK_HINTS = (
    "contact",
    "contacte",
    "contact us",
    "contactati ne",
    "contactați ne",
    "ia legatura",
    "ia legătura",
    "get in touch",
    "reach us",
)

KNOWLEDGE_WEBSITE_HINTS = (
    "site",
    "website",
    "official website",
    "visit website",
    "visit official site",
    "site web",
    "site oficial",
    "pagina web",
    "pagină web",
    "business website",
    "location website",
)

FOOTER_SELECTORS = (
    "footer",
    "[role='contentinfo']",
    "#footer",
    ".footer",
    "[class*='footer']",
    "[id*='footer']",
)

STREET_STOPWORDS = {
    "str",
    "strada",
    "bd",
    "bulevard",
    "calea",
    "nr",
    "bloc",
    "bl",
    "ap",
    "apartament",
    "parter",
    "etaj",
    "loc",
    "localitatea",
    "sat",
    "comuna",
    "municipiul",
    "jud",
    "judet",
    "județ",
}

EMAIL_RE = re.compile(
    r"(?i)(?<![\w.+-])"
    r"[a-z0-9._%+-]{1,64}@[a-z0-9.-]{1,190}\.[a-z]{2,24}"
    r"(?![\w.-])"
)

PHONE_RE = re.compile(
    r"(?<!\d)"
    r"(?:\+40|0040|0)"
    r"(?:[\s()./-]*\d){9}"
    r"(?!\d)"
)

OBFUSCATED_EMAIL_RE = re.compile(
    r"(?i)\b[a-z0-9._%+-]{1,64}\s*"
    r"(?:@|\[\s*at\s*\]|\(\s*at\s*\)|\s+at\s+)\s*"
    r"[a-z0-9-]+(?:\s*(?:\.|\[\s*dot\s*\]|\(\s*dot\s*\)|\s+dot\s+)"
    r"\s*[a-z0-9-]+)+\b"
)

CUI_LABEL_RE = re.compile(
    r"(?i)\b(?:c\.?\s*u\.?\s*i\.?|c\.?\s*i\.?\s*f\.?|"
    r"cod(?:ul)?\s+(?:unic\s+de\s+[iî]nregistrare|fiscal))"
    r"\s*(?:nr\.?|numar(?:ul)?)?\s*[:#-]?\s*(?:ro\s*)?(\d{2,10})\b"
)


@dataclass
class GoogleCandidate:
    url: str
    domain: str
    title: str
    snippet: str
    query: str
    rank: int
    score: int
    domain_score: int
    company_score: int
    location_score: int
    cui_found: bool
    source: str
    accepted: bool
    rejection_reason: str = ""
    panel_context: str = ""
    duplicate_cuis: list[str] = field(default_factory=list)


@dataclass
class ContactValue:
    value: str
    source: str
    score: int


@dataclass
class ContactResult:
    website: str = ""
    google_source: str = ""
    google_rank: int = 0
    google_title: str = ""
    google_query: str = ""
    panel_context: str = ""
    website_cuis: list[str] = field(default_factory=list)
    cui_match_status: str = ""
    contact_page_url: str = ""
    email: str = ""
    phone: str = ""
    all_emails: list[str] = field(default_factory=list)
    all_phones: list[str] = field(default_factory=list)
    found_in: str = ""
    status: str = "NO_WEBSITE"
    website_score: int = 0
    duplicate_cuis: list[str] = field(default_factory=list)
    checked_at: str = ""
    notes: str = ""


def merge_google_panel_contact(
    result: ContactResult,
    panel_contact: ContactResult,
) -> ContactResult:
    if not result.status.startswith("FOUND_"):
        panel_contact.notes += f"; linked-site result was {result.status}"
        return panel_contact

    result.all_emails = list(
        dict.fromkeys(result.all_emails + panel_contact.all_emails)
    )
    result.all_phones = list(
        dict.fromkeys(result.all_phones + panel_contact.all_phones)
    )
    result.email = result.email or panel_contact.email
    result.phone = result.phone or panel_contact.phone
    if panel_contact.found_in and panel_contact.found_in not in result.found_in:
        result.found_in = "; ".join(
            item for item in (result.found_in, panel_contact.found_in) if item
        )
    result.notes += "; matching Google-panel contacts included"
    return result


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def normalize_header(value: Any) -> str:
    return normalize_text(value).replace(" ", "_")


def normalize_cui(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\D+", "", text)


def collapse_initial_tokens(value: str) -> list[str]:
    result: list[str] = []
    initials: list[str] = []

    def flush_initials() -> None:
        if initials:
            result.append("".join(initials))
            initials.clear()

    for token in normalize_text(value).split():
        if len(token) == 1 and token.isalpha():
            initials.append(token)
        else:
            flush_initials()
            result.append(token)
    flush_initials()
    return result


def company_tokens(company: str) -> list[str]:
    return [
        token
        for token in collapse_initial_tokens(clean_company_for_search(company))
        if len(token) >= 2 and token not in LEGAL_SUFFIXES
    ]


def clean_company_for_search(company: str) -> str:
    cleaned = re.sub(r"\s+", " ", company).strip()
    cleaned = re.sub(r"\s*&\s*", "&", cleaned)
    cleaned = re.sub(r"(?i)^s\.?\s*c\.?\s+", "", cleaned)
    cleaned = re.sub(
        r"(?i)\s+(?:s\.?\s*r\.?\s*l\.?|s\.?\s*a\.?)$",
        "",
        cleaned,
    )
    return cleaned.strip(" ,")


def clean_address_for_search(address: str) -> str:
    cleaned = re.sub(r"(?i)\b(?:nr|numar(?:ul)?)\.?\s*[:#-]?\s*", "", address)
    cleaned = re.sub(r"[,;]+", " ", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def significant_company_tokens(company: str) -> list[str]:
    return [
        token
        for token in company_tokens(company)
        if len(token) >= 3 and token not in GENERIC_DOMAIN_WORDS
    ]


def canonical_domain(url: str) -> str:
    try:
        host = (urlparse(url).hostname or "").casefold().strip(".")
    except ValueError:
        return ""
    if host.startswith("www."):
        host = host[4:]
    return host


def unwrap_google_url(url: str) -> str:
    """Extract the real target from common Google redirect URLs."""
    url = (url or "").strip()
    if not url:
        return ""

    if url.startswith("/"):
        url = urljoin("https://www.google.com", url)

    try:
        parsed = urlparse(url)
    except ValueError:
        return url

    host = (parsed.hostname or "").casefold()
    if host.endswith("google.com") or host.endswith("google.ro"):
        params = parse_qs(parsed.query)
        for key in ("url", "q"):
            values = params.get(key)
            if values and values[0].startswith(("http://", "https://")):
                return values[0]

    return url


def normalize_url(url: str) -> str:
    url = unwrap_google_url(url)
    if not url:
        return ""
    if url.startswith(("mailto:", "tel:", "javascript:", "#")):
        return ""
    if url.startswith("//"):
        url = "https:" + url
    if not re.match(r"(?i)^https?://", url):
        return ""

    try:
        parsed = urlparse(url)
    except ValueError:
        return ""

    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""

    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path or "/",
            "",
            parsed.query,
            "",
        )
    )


def base_url(url: str) -> str:
    parsed = urlparse(url)
    return urlunparse((parsed.scheme, parsed.netloc, "/", "", "", ""))


def is_excluded_domain(domain: str) -> bool:
    return any(
        domain == item or domain.endswith("." + item)
        for item in EXCLUDED_DOMAINS
    )


def is_verified_contact_directory(domain: str) -> bool:
    return any(
        domain == item or domain.endswith("." + item)
        for item in VERIFIED_CONTACT_DIRECTORIES
    )


def is_google_domain(domain: str) -> bool:
    return any(
        domain == item or domain.endswith("." + item)
        for item in ("google.com", "google.ro", "googleusercontent.com")
    )


def is_verified_listing(candidate: GoogleCandidate) -> bool:
    return (
        candidate.source in {"knowledge_panel", "organic_listing"}
        and is_verified_contact_directory(candidate.domain)
        and not is_google_domain(candidate.domain)
    )


def company_match_score(company: str, text: str) -> int:
    tokens = company_tokens(company)
    if not tokens:
        return 0

    haystack = set(normalize_text(text).split())
    haystack.update(collapse_initial_tokens(text))
    matched = sum(token in haystack for token in tokens)
    ratio = matched / len(tokens)

    if ratio >= 1.0:
        return 100
    if ratio >= 0.75:
        return 85
    if ratio >= 0.50:
        return 65
    if ratio > 0:
        return 35
    return 0


def domain_company_score(company: str, domain: str) -> int:
    tokens = significant_company_tokens(company)
    if not tokens:
        return 0

    stem = normalize_text(domain.split(".", 1)[0]).replace(" ", "")
    joined = "".join(tokens)
    matched = sum(token.replace(" ", "") in stem for token in tokens)
    ratio = matched / len(tokens)

    if joined and joined in stem:
        return 50
    if ratio >= 1.0:
        return 45
    if ratio >= 0.50:
        return 32
    if ratio > 0:
        return 22
    return 0


def address_tokens(address: str) -> list[str]:
    result: list[str] = []
    for token in normalize_text(address).split():
        if token in STREET_STOPWORDS or token.isdigit() or len(token) < 4:
            continue
        if token not in result:
            result.append(token)
    return result[:5]


def address_street_number(address: str) -> str:
    match = re.search(
        r"(?i)\b(?:nr|numar(?:ul)?)\.?\s*[:#-]?\s*(\d+[a-z]?)\b",
        address,
    )
    return normalize_text(match.group(1)) if match else ""


def address_identity_match(county: str, address: str, text: str) -> bool:
    normalized_text = normalize_text(text)
    haystack = set(normalized_text.split())
    matched_tokens = sum(token in haystack for token in address_tokens(address))
    street_number = address_street_number(address)
    number_found = bool(street_number and street_number in haystack)
    county_normalized = normalize_text(county)
    county_found = bool(county_normalized and county_normalized in normalized_text)

    if street_number:
        return (
            number_found
            and matched_tokens >= 1
            and (county_found or matched_tokens >= 2)
        )
    return matched_tokens >= 2 and county_found


def location_match_score(county: str, address: str, text: str) -> int:
    normalized = normalize_text(text)
    score = 0

    county_normalized = normalize_text(county)
    if county_normalized and county_normalized in normalized:
        score += 30

    matched_address = sum(
        token in normalized for token in address_tokens(address)
    )
    if matched_address >= 2:
        score += 20
    elif matched_address == 1:
        score += 10

    return min(50, score)


def contains_cui(cui: str, text: str) -> bool:
    if not cui:
        return False
    return bool(re.search(rf"(?<!\d){re.escape(cui)}(?!\d)", text))


def extract_labeled_cuis(text: str) -> list[str]:
    """Extract fiscal identifiers only when they have an explicit CUI/CIF label."""
    return list(dict.fromkeys(match.group(1) for match in CUI_LABEL_RE.finditer(text)))


def identity_decision(
    company: str,
    cui: str,
    county: str,
    address: str,
    text: str,
) -> tuple[bool, list[str], str, str]:
    website_cuis = extract_labeled_cuis(text)
    if website_cuis:
        if cui in website_cuis:
            return True, website_cuis, "MATCH", "CUI matched"
        return False, website_cuis, "MISMATCH", "Published CUI belongs to another company"

    company_score = company_match_score(company, text)
    if company_score >= 85 and address_identity_match(county, address, text):
        return (
            True,
            [],
            "NOT_FOUND",
            "CUI not published; company name and street address matched",
        )
    return (
        False,
        [],
        "NOT_FOUND",
        "CUI not published and company name/address evidence was insufficient",
    )


def is_ambiguous_company(company: str, name_frequency: Counter[str]) -> bool:
    normalized = normalize_text(company)
    tokens = company_tokens(company)

    if name_frequency[normalized] > 1:
        return True
    if not significant_company_tokens(company):
        return True
    if len(tokens) == 1 and tokens[0] in GENERIC_DOMAIN_WORDS:
        return True
    return False


def duplicate_cuis_for_domain(
    domain: str,
    current_cui: str,
    assigned_domains: dict[str, set[str]],
) -> list[str]:
    return sorted(
        cui for cui in assigned_domains.get(domain, set())
        if cui != current_cui
    )


def candidate_decision(
    *,
    company: str,
    cui: str,
    county: str,
    address: str,
    url: str,
    title: str,
    context: str,
    query: str,
    rank: int,
    source: str,
    min_website_score: int,
    name_frequency: Counter[str],
    assigned_domains: dict[str, set[str]],
) -> GoogleCandidate:
    domain = canonical_domain(url)
    combined = f"{title}\n{context}"
    company_score = company_match_score(company, combined)
    domain_score = domain_company_score(company, domain)
    location_score = location_match_score(county, address, combined)
    cui_found = contains_cui(cui, combined)
    duplicate_cuis = duplicate_cuis_for_domain(
        domain=domain,
        current_cui=cui,
        assigned_domains=assigned_domains,
    )

    rank_score = 18 if source == "knowledge_panel" else max(0, 12 - rank)
    source_bonus = 35 if source == "knowledge_panel" else 0
    cui_bonus = 35 if cui_found else 0
    score = min(
        100,
        source_bonus
        + rank_score
        + domain_score
        + min(25, company_score // 4)
        + min(20, location_score)
        + cui_bonus,
    )

    ambiguous = is_ambiguous_company(company, name_frequency)
    accepted = False
    reason = ""
    verified_listing = (
        source in {"knowledge_panel", "organic_listing"}
        and is_verified_contact_directory(domain)
        and not is_google_domain(domain)
    )

    if not domain or is_google_domain(domain):
        reason = "excluded or invalid domain"
    elif is_excluded_domain(domain) and not verified_listing:
        reason = "excluded domain"
    elif duplicate_cuis and not verified_listing:
        reason = "domain already assigned to another CUI"
    elif source == "knowledge_panel":
        if cui_found and company_score >= 65:
            accepted = True
        elif company_score >= 85 and address_identity_match(
            county,
            address,
            combined,
        ):
            accepted = True
        else:
            reason = "knowledge panel lacks matching CUI or full name/address"
    elif source == "organic_listing":
        if cui_found and company_score >= 65:
            accepted = True
        elif company_score >= 85 and address_identity_match(
            county,
            address,
            combined,
        ):
            accepted = True
        else:
            reason = "directory listing lacks matching CUI or full name/address"
    elif company_score < 65:
        reason = "company name does not match Google context"
    elif ambiguous:
        # Never accept APICOLA/MATCA-like organic matches from name/domain alone.
        if cui_found and domain_score >= 22 and score >= min_website_score:
            accepted = True
        elif location_score >= 20 and domain_score >= 22 and score >= min_website_score:
            accepted = True
        else:
            reason = "ambiguous company name without location/CUI evidence"
    elif cui_found and domain_score >= 22 and score >= min_website_score:
        accepted = True
    elif domain_score >= 22 and score >= min_website_score:
        accepted = True
    else:
        reason = "insufficient official-website evidence"

    return GoogleCandidate(
        url=url,
        domain=domain,
        title=title,
        snippet=context,
        query=query,
        rank=rank,
        score=score,
        domain_score=domain_score,
        company_score=company_score,
        location_score=location_score,
        cui_found=cui_found,
        source=source,
        accepted=accepted,
        rejection_reason=reason,
        panel_context=context[:1000] if source == "knowledge_panel" else "",
        duplicate_cuis=duplicate_cuis,
    )


def build_google_queries(
    company: str,
    cui: str,
    county: str,
    address: str,
    ambiguous: bool,
) -> list[str]:
    clean_company = clean_company_for_search(company)
    county_clean = re.sub(r"\s+", " ", county).strip()
    clean_address = clean_address_for_search(address)

    identity_phrase = " ".join(
        part for part in (clean_company, clean_address, county_clean) if part
    )
    first_query = f'"{identity_phrase}"'

    exact_cui_query = f'"{clean_company}" "{cui}"'
    contact_terms = "(contact OR telefon OR email)"
    cui_contact_query = f'"{cui}" {contact_terms}'
    company_contact_query = " ".join(
        part
        for part in (f'"{clean_company}"', county_clean, contact_terms)
        if part
    )

    result: list[str] = []
    queries = [first_query, exact_cui_query]
    if ambiguous:
        queries.extend((cui_contact_query, company_contact_query))
    for query in queries:
        if query and query not in result:
            result.append(query)
    return result


def looks_like_google_block(page: Page) -> bool:
    url = page.url.casefold()
    try:
        title = normalize_text(page.title())
    except Exception:
        title = ""
    try:
        text = normalize_text(page.locator("body").inner_text(timeout=5_000))
    except Exception:
        text = ""

    indicators = (
        "unusual traffic",
        "trafic neobisnuit",
        "verify you are human",
        "verificati ca sunteti om",
        "captcha",
        "before you continue to google",
        "inainte de a continua la google",
    )
    return "/sorry/" in url or any(
        item in title or item in text for item in indicators
    )


def wait_for_google_content(page: Page, manual_captcha: bool) -> bool:
    ready = page.locator("#search, #rhs")
    try:
        ready.first.wait_for(state="attached", timeout=12_000)
        return True
    except PlaywrightTimeoutError:
        pass

    if not manual_captcha:
        return False

    print()
    if looks_like_google_block(page):
        print("  Google requires manual consent/CAPTCHA verification.")
    else:
        print("  Google results or company panel are not visible yet.")
    input("  Resolve the page in Chromium, then press Enter here... ")

    try:
        ready.first.wait_for(state="attached", timeout=20_000)
        return True
    except PlaywrightTimeoutError:
        return False


def safe_inner_text(locator: Locator, timeout: int = 3_000) -> str:
    try:
        return locator.inner_text(timeout=timeout).strip()
    except Exception:
        return ""


def anchor_metadata(anchor: Locator) -> str:
    try:
        return str(
            anchor.evaluate(
                """
                (a) => {
                    const parts = [];
                    if (a.textContent) parts.push(a.textContent);
                    for (const child of a.querySelectorAll(
                        '[aria-label], [title], [data-item-id], [data-attrid]'
                    )) {
                        for (const name of ['aria-label', 'title', 'data-item-id', 'data-attrid']) {
                            const value = child.getAttribute(name);
                            if (value) parts.push(value);
                        }
                    }
                    let node = a;
                    for (let depth = 0; depth < 6 && node; depth += 1) {
                        for (const name of ['data-attrid', 'aria-label', 'title', 'class', 'id']) {
                            const value = node.getAttribute && node.getAttribute(name);
                            if (value) parts.push(value);
                        }
                        if (depth === 0 && node.innerText) parts.push(node.innerText);
                        node = node.parentElement;
                    }
                    return parts.join(' ');
                }
                """
            )
            or ""
        )
    except Exception:
        return ""


def has_knowledge_website_hint(text: str) -> bool:
    normalized = normalize_text(text)
    tokens = set(normalized.split())
    return any(
        hint == "site" and "site" in tokens
        or hint != "site" and normalize_text(hint) in normalized
        for hint in KNOWLEDGE_WEBSITE_HINTS
    )


def knowledge_panel_context(anchor: Locator) -> str:
    """Return nearby panel text without falling back to the whole results page."""
    try:
        return str(
            anchor.evaluate(
                """
                (a) => {
                    let node = a;
                    let best = '';
                    for (let depth = 0; depth < 12 && node; depth += 1) {
                        const text = (node.innerText || '').trim();
                        if (text.length >= 40 && text.length <= 4000) best = text;
                        if (
                            node.id === 'rhs' ||
                            (node.getAttribute && node.getAttribute('role') === 'complementary')
                        ) return text;
                        if (node.id === 'search' || node.tagName === 'BODY') break;
                        node = node.parentElement;
                    }
                    return best;
                }
                """
            )
            or ""
        )
    except Exception:
        return ""


def google_panel_text(page: Page) -> str:
    selectors = (
        "#rhs",
        "[role='complementary']",
        "[data-attrid*='kc:/']",
    )
    best = ""
    for selector in selectors:
        locator = page.locator(selector)
        try:
            for index in range(min(locator.count(), 20)):
                item = locator.nth(index)
                if not item.is_visible():
                    continue
                text = safe_inner_text(item, timeout=3_000)
                if len(text) > len(best):
                    best = text
        except Exception:
            continue
    return best


def extract_google_panel_contact(
    *,
    page: Page,
    company: str,
    cui: str,
    county: str,
    address: str,
    query: str,
) -> ContactResult | None:
    panel_text = google_panel_text(page)
    if not panel_text:
        return None

    identity_ok, panel_cuis, cui_match_status, identity_notes = identity_decision(
        company,
        cui,
        county,
        address,
        panel_text,
    )
    if not identity_ok:
        return None

    emails = merge_contact_values(
        collect_emails_from_text(panel_text, "google_panel", "")
    )
    phones = merge_contact_values(
        collect_phones_from_text(panel_text, "google_panel")
    )
    if not emails and not phones:
        return None

    return ContactResult(
        google_source="knowledge_panel",
        google_title=company,
        google_query=query,
        panel_context=panel_text[:1000],
        website_cuis=panel_cuis,
        cui_match_status=cui_match_status,
        email=emails[0].value if emails else "",
        phone=phones[0].value if phones else "",
        all_emails=[item.value for item in emails],
        all_phones=[item.value for item in phones],
        found_in="google_panel",
        status="FOUND_GOOGLE_PANEL",
        checked_at=datetime.now().isoformat(timespec="seconds"),
        notes=f"Google panel identity verified; {identity_notes}",
    )


def extract_knowledge_panel_candidate(
    *,
    page: Page,
    company: str,
    cui: str,
    county: str,
    address: str,
    query: str,
    min_website_score: int,
    name_frequency: Counter[str],
    assigned_domains: dict[str, set[str]],
) -> GoogleCandidate | None:
    # Google moves the visible Site/Website label between the link and nested
    # button elements. Cover both layouts as well as the traditional #rhs.
    anchors = page.locator(
        '#rhs a[href], '
        '[role="complementary"] a[href], '
        '[data-attrid*="website"] a[href], '
        '[data-attrid*="official_site"] a[href], '
        '[data-item-id*="authority"] a[href], '
        'a[href]:has([data-item-id*="authority"]), '
        'a[href][data-item-id*="authority"], '
        '[aria-label="site" i] a[href], '
        '[aria-label*="website" i] a[href], '
        'a[href]:has([aria-label="site" i]), '
        'a[href]:has([aria-label*="website" i]), '
        'a[href][aria-label*="website" i], '
        'a[href][aria-label="site" i], '
        'a[href][aria-label*="site web" i], '
        'a[href][aria-label*="site oficial" i]'
    )

    best: GoogleCandidate | None = None
    seen_urls: set[str] = set()

    try:
        count = min(anchors.count(), 160)
    except Exception:
        count = 0

    for index in range(count):
        anchor = anchors.nth(index)
        href = normalize_url(anchor.get_attribute("href") or "")
        if not href or href in seen_urls:
            continue
        seen_urls.add(href)

        domain = canonical_domain(href)
        # Exact Site links from the company panel may legitimately point to a
        # directory listing. Organic candidates keep the exclusion rule.
        if (
            not domain
            or is_google_domain(domain)
            or (
                is_excluded_domain(domain)
                and not is_verified_contact_directory(domain)
            )
        ):
            continue

        label = safe_inner_text(anchor, timeout=1_500)
        metadata = anchor_metadata(anchor)
        hint_text = normalize_text(f"{label} {metadata}")

        if not has_knowledge_website_hint(hint_text):
            continue

        panel_text = knowledge_panel_context(anchor)
        candidate = candidate_decision(
            company=company,
            cui=cui,
            county=county,
            address=address,
            url=href,
            title=label or domain,
            context=panel_text,
            query=query,
            rank=0,
            source="knowledge_panel",
            min_website_score=min_website_score,
            name_frequency=name_frequency,
            assigned_domains=assigned_domains,
        )

        if best is None or candidate.score > best.score:
            best = candidate

    return best


def result_snippet(link: Locator) -> str:
    try:
        return str(
            link.evaluate(
                """
                (anchor) => {
                    let node = anchor;
                    for (let depth = 0; depth < 7 && node; depth += 1) {
                        const text = (node.innerText || '').trim();
                        if (text.length >= 60 && text.length <= 1800) return text;
                        node = node.parentElement;
                    }
                    return '';
                }
                """
            )
            or ""
        )
    except Exception:
        return ""


def extract_organic_candidates(
    *,
    page: Page,
    company: str,
    cui: str,
    county: str,
    address: str,
    query: str,
    max_results: int,
    min_website_score: int,
    name_frequency: Counter[str],
    assigned_domains: dict[str, set[str]],
) -> list[GoogleCandidate]:
    candidates: list[GoogleCandidate] = []
    seen_domains: set[str] = set()
    links = page.locator("a:has(h3)")

    try:
        count = min(links.count(), max_results)
    except Exception:
        count = 0

    for index in range(count):
        link = links.nth(index)
        href = normalize_url(link.get_attribute("href") or "")
        if not href:
            continue

        domain = canonical_domain(href)
        if not domain or domain in seen_domains or is_google_domain(domain):
            continue
        if is_excluded_domain(domain) and not is_verified_contact_directory(domain):
            continue

        title_locator = link.locator("h3")
        if title_locator.count() == 0:
            continue

        title = safe_inner_text(title_locator.first)
        snippet = result_snippet(link)
        source = (
            "organic_listing"
            if is_verified_contact_directory(domain)
            else "organic_result"
        )
        candidate = candidate_decision(
            company=company,
            cui=cui,
            county=county,
            address=address,
            url=href,
            title=title,
            context=snippet,
            query=query,
            rank=index + 1,
            source=source,
            min_website_score=min_website_score,
            name_frequency=name_frequency,
            assigned_domains=assigned_domains,
        )
        if source == "organic_listing" and not candidate.accepted:
            continue
        candidates.append(candidate)
        seen_domains.add(domain)

    return sorted(
        candidates,
        key=lambda item: (-item.accepted, -item.score, item.rank),
    )


def discover_website(
    *,
    page: Page,
    company: str,
    cui: str,
    county: str,
    address: str,
    max_results: int,
    min_website_score: int,
    manual_captcha: bool,
    google_delay: float,
    name_frequency: Counter[str],
    assigned_domains: dict[str, set[str]],
) -> tuple[
    GoogleCandidate | None,
    GoogleCandidate | None,
    bool,
    ContactResult | None,
]:
    best_review: GoogleCandidate | None = None
    google_blocked = False
    ambiguous = is_ambiguous_company(company, name_frequency)

    queries = build_google_queries(
        company=company,
        cui=cui,
        county=county,
        address=address,
        ambiguous=ambiguous,
    )

    for query_number, query in enumerate(queries, start=1):
        print(f"  Google query {query_number}: {query}")
        search_url = (
            "https://www.google.com/search?"
            f"q={quote_plus(query)}&num={max_results}"
        )

        try:
            page.goto(
                search_url,
                wait_until="domcontentloaded",
                timeout=60_000,
            )
        except PlaywrightTimeoutError:
            print("    Google navigation timed out.")
            continue

        if not wait_for_google_content(page, manual_captcha):
            if looks_like_google_block(page):
                google_blocked = True
            continue

        # IMPORTANT: inspect the right-side company panel before organic results.
        panel_contact = extract_google_panel_contact(
            page=page,
            company=company,
            cui=cui,
            county=county,
            address=address,
            query=query,
        )
        if panel_contact is not None:
            print(
                "    Knowledge panel contact ACCEPT: "
                f"phone={panel_contact.phone or '-'}, "
                f"email={panel_contact.email or '-'}"
            )

        panel_candidate = extract_knowledge_panel_candidate(
            page=page,
            company=company,
            cui=cui,
            county=county,
            address=address,
            query=query,
            min_website_score=min_website_score,
            name_frequency=name_frequency,
            assigned_domains=assigned_domains,
        )

        if panel_candidate is not None:
            marker = "ACCEPT" if panel_candidate.accepted else "review"
            print(
                f"    Knowledge panel {marker}: "
                f"domain={panel_candidate.domain}, "
                f"score={panel_candidate.score}, "
                f"location={panel_candidate.location_score}"
            )
            if panel_candidate.accepted:
                return panel_candidate, best_review, google_blocked, panel_contact
            if best_review is None or panel_candidate.score > best_review.score:
                best_review = panel_candidate

        if panel_contact is not None:
            return None, best_review, google_blocked, panel_contact

        candidates = extract_organic_candidates(
            page=page,
            company=company,
            cui=cui,
            county=county,
            address=address,
            query=query,
            max_results=max_results,
            min_website_score=min_website_score,
            name_frequency=name_frequency,
            assigned_domains=assigned_domains,
        )

        for candidate in candidates[:5]:
            marker = "ACCEPT" if candidate.accepted else "review"
            reason = f" ({candidate.rejection_reason})" if not candidate.accepted else ""
            result_kind = (
                "Verified listing"
                if candidate.source == "organic_listing"
                else "Organic"
            )
            print(
                f"    {result_kind} {marker}: rank={candidate.rank}, "
                f"domain={candidate.domain}, score={candidate.score}{reason}"
            )

        accepted = next(
            (candidate for candidate in candidates if candidate.accepted),
            None,
        )
        if accepted is not None:
            return accepted, best_review, google_blocked, None

        if candidates:
            candidate = candidates[0]
            if best_review is None or candidate.score > best_review.score:
                best_review = candidate

        if query_number < len(queries):
            time.sleep(max(0.0, google_delay))

    return None, best_review, google_blocked, None


def normalize_phone(raw: str) -> str:
    digits = re.sub(r"\D+", "", raw)
    if digits.startswith("400") and len(digits) == 12:
        digits = digits[3:]
    if digits.startswith("0040"):
        digits = digits[4:]
    elif digits.startswith("40") and len(digits) == 11:
        digits = digits[2:]
    elif digits.startswith("0"):
        digits = digits[1:]
    if len(digits) != 9:
        return ""
    return "+40" + digits


def valid_email(email: str) -> bool:
    email = email.strip(" .,:;<>[](){}").casefold()
    if not EMAIL_RE.fullmatch(email):
        return False
    local, domain = email.rsplit("@", 1)
    if domain in TECHNICAL_EMAIL_DOMAINS:
        return False
    if any(
        email.endswith(extension)
        for extension in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg")
    ):
        return False
    return bool(local and "." in domain)


def normalize_obfuscated_email(raw: str) -> str:
    value = re.sub(r"(?i)\s*(?:\[\s*at\s*\]|\(\s*at\s*\)|\s+at\s+)\s*", "@", raw)
    value = re.sub(r"(?i)\s*(?:\[\s*dot\s*\]|\(\s*dot\s*\)|\s+dot\s+)\s*", ".", value)
    return re.sub(r"\s+", "", value).casefold()


def decode_cloudflare_email(encoded: str) -> str:
    try:
        key = int(encoded[:2], 16)
        return "".join(
            chr(int(encoded[index:index + 2], 16) ^ key)
            for index in range(2, len(encoded), 2)
        ).casefold()
    except (TypeError, ValueError):
        return ""


def email_score(email: str, source: str, website_domain: str) -> int:
    local, domain = email.rsplit("@", 1)
    score = 0
    if domain == website_domain or domain.endswith("." + website_domain):
        score += 45
    if local.startswith(GENERIC_EMAIL_PREFIXES):
        score += 35
    if local.startswith(LOW_PRIORITY_EMAIL_PREFIXES):
        score -= 40
    if source == "contact_page":
        score += 25
    elif source == "footer":
        score += 15
    return score


def phone_score(phone: str, source: str, context: str = "") -> int:
    score = 25 if source == "contact_page" else 15
    if phone.startswith(("+402", "+403")):
        score += 8
    elif phone.startswith("+407"):
        score += 5
    normalized_context = normalize_text(context)
    if "fax" in normalized_context:
        score -= 35
    elif any(label in normalized_context for label in ("tel", "telefon", "phone")):
        score += 10
    return score


def collect_emails_from_text(
    text: str,
    source: str,
    website_domain: str,
) -> list[ContactValue]:
    values: dict[str, ContactValue] = {}
    candidates = [match.group(0) for match in EMAIL_RE.finditer(text)]
    candidates.extend(
        normalize_obfuscated_email(match.group(0))
        for match in OBFUSCATED_EMAIL_RE.finditer(text)
    )
    for raw_email in candidates:
        email = raw_email.strip(" .,:;<>[](){}").casefold()
        if not valid_email(email):
            continue
        values[email] = ContactValue(
            value=email,
            source=source,
            score=email_score(email, source, website_domain),
        )
    return list(values.values())


def collect_emails(
    text: str,
    locator: Locator,
    source: str,
    website_domain: str,
) -> list[ContactValue]:
    values = {
        item.value: item
        for item in collect_emails_from_text(text, source, website_domain)
    }
    try:
        links = locator.locator('a[href^="mailto:"]')
        for index in range(min(links.count(), 30)):
            href = links.nth(index).get_attribute("href") or ""
            address = href[7:].split("?", 1)[0].strip().casefold()
            for email in re.split(r"[,;]", address):
                if not valid_email(email):
                    continue
                values[email] = ContactValue(
                    value=email,
                    source=source,
                    score=email_score(email, source, website_domain) + 5,
                )
    except Exception:
        pass
    try:
        protected = locator.locator("[data-cfemail]")
        for index in range(min(protected.count(), 30)):
            encoded = protected.nth(index).get_attribute("data-cfemail") or ""
            email = decode_cloudflare_email(encoded)
            if not valid_email(email):
                continue
            values[email] = ContactValue(
                value=email,
                source=source,
                score=email_score(email, source, website_domain) + 5,
            )
    except Exception:
        pass
    return list(values.values())


def collect_phones_from_text(text: str, source: str) -> list[ContactValue]:
    values: dict[str, ContactValue] = {}
    for index, match in enumerate(PHONE_RE.finditer(text)):
        phone = normalize_phone(match.group(0))
        if not phone:
            continue
        nearby_text = text[max(0, match.start() - 40):match.start()]
        values[phone] = ContactValue(
            value=phone,
            source=source,
            score=phone_score(phone, source, nearby_text) + max(0, 10 - index),
        )
    return list(values.values())


def collect_phones(
    text: str,
    locator: Locator,
    source: str,
) -> list[ContactValue]:
    values = {
        item.value: item for item in collect_phones_from_text(text, source)
    }
    try:
        links = locator.locator('a[href^="tel:"]')
        for index in range(min(links.count(), 30)):
            href = links.nth(index).get_attribute("href") or ""
            phone = normalize_phone(href[4:])
            if not phone:
                continue
            candidate = ContactValue(
                value=phone,
                source=source,
                score=(
                    phone_score(
                        phone,
                        source,
                        safe_inner_text(links.nth(index), timeout=1_000),
                    )
                    + 5
                ),
            )
            existing = values.get(phone)
            if existing is None or candidate.score > existing.score:
                values[phone] = candidate
    except Exception:
        pass
    return list(values.values())


def footer_locator(page: Page) -> Locator | None:
    best: Locator | None = None
    best_length = -1
    for selector in FOOTER_SELECTORS:
        locator = page.locator(selector)
        try:
            for index in range(min(locator.count(), 10)):
                item = locator.nth(index)
                if not item.is_visible():
                    continue
                length = len(safe_inner_text(item, timeout=2_000))
                if length > best_length:
                    best = item
                    best_length = length
        except Exception:
            continue
    return best


def footer_data(
    page: Page,
    website_domain: str,
) -> tuple[list[ContactValue], list[ContactValue], list[str]]:
    locator = footer_locator(page)
    if locator is not None:
        try:
            text = locator.inner_text(timeout=10_000)
            return (
                collect_emails(text, locator, "footer", website_domain),
                collect_phones(text, locator, "footer"),
                extract_labeled_cuis(text),
            )
        except Exception:
            pass

    # Fallback: only the last visible 5,000 characters, not the whole homepage.
    try:
        body = page.locator("body")
        bottom_text = body.inner_text(timeout=10_000)[-5000:]
        return (
            collect_emails_from_text(bottom_text, "footer", website_domain),
            collect_phones_from_text(bottom_text, "footer"),
            extract_labeled_cuis(bottom_text),
        )
    except Exception:
        return [], [], []


def contact_link_score(text: str, href: str) -> int:
    normalized_text = normalize_text(text)
    path = normalize_text(urlparse(href).path)
    combined = f"{normalized_text} {path}"
    score = 0
    if normalized_text in {"contact", "contacte", "contact us"}:
        score += 60
    if any(hint in combined for hint in CONTACT_LINK_HINTS):
        score += 35
    if "/contact" in urlparse(href).path.casefold():
        score += 25
    if any(word in combined for word in ("blog", "news", "stiri", "articol", "article")):
        score -= 25
    return score


def discover_contact_url(
    page: Page,
    website_domain: str,
    google_result_url: str,
) -> str:
    candidates: dict[str, int] = {}
    google_url = normalize_url(google_result_url)
    if (
        google_url
        and canonical_domain(google_url) == website_domain
        and "contact" in normalize_text(urlparse(google_url).path)
    ):
        candidates[google_url] = 100

    anchors = page.locator("a[href]")
    try:
        count = min(anchors.count(), 500)
    except Exception:
        count = 0

    for index in range(count):
        anchor = anchors.nth(index)
        try:
            raw_href = anchor.get_attribute("href") or ""
        except Exception:
            continue
        if raw_href.startswith(("mailto:", "tel:", "javascript:", "#")):
            continue

        href = normalize_url(urljoin(page.url, raw_href))
        if not href or canonical_domain(href) != website_domain:
            continue

        text = safe_inner_text(anchor, timeout=1_500)
        score = contact_link_score(text, href)
        if score > 0:
            candidates[href] = max(score, candidates.get(href, 0))

    return max(candidates, key=candidates.get) if candidates else ""


def open_page(page: Page, url: str) -> bool:
    try:
        response = page.goto(
            url,
            wait_until="domcontentloaded",
            timeout=45_000,
        )
        return not response or response.status < 400
    except Exception:
        return False


def contact_page_data(
    page: Page,
    contact_url: str,
    website_domain: str,
) -> tuple[list[ContactValue], list[ContactValue], list[str], str, str]:
    if not contact_url or not open_page(page, contact_url):
        return [], [], [], "", ""
    try:
        page.wait_for_load_state("networkidle", timeout=5_000)
    except Exception:
        pass
    try:
        body = page.locator("body")
        text = body.inner_text(timeout=15_000)
    except Exception:
        return [], [], [], page.url, ""
    return (
        collect_emails(text, body, "contact_page", website_domain),
        collect_phones(text, body, "contact_page"),
        extract_labeled_cuis(text),
        page.url,
        text,
    )


def primary_content_locator(page: Page) -> Locator:
    selectors = (
        "[itemtype*='LocalBusiness']",
        "[class*='listing-detail']",
        "[class*='company-detail']",
        "main",
        "[role='main']",
        "#main",
        "#content",
    )
    for selector in selectors:
        locator = page.locator(selector)
        best: Locator | None = None
        best_length = -1
        try:
            for index in range(min(locator.count(), 12)):
                item = locator.nth(index)
                if not item.is_visible():
                    continue
                length = len(safe_inner_text(item, timeout=2_000))
                if length > best_length:
                    best = item
                    best_length = length
        except Exception:
            continue
        if best is not None and best_length >= 20:
            return best
    return page.locator("body")


def inspect_verified_listing(
    page: Page,
    candidate: GoogleCandidate,
    expected_company: str,
    expected_cui: str,
    expected_county: str,
    expected_address: str,
) -> ContactResult:
    checked_at = datetime.now().isoformat(timespec="seconds")
    print(f"  Opening exact verified listing: {candidate.url}")
    if not open_page(page, candidate.url):
        return ContactResult(
            website=candidate.url,
            google_source=candidate.source,
            google_title=candidate.title,
            google_query=candidate.query,
            panel_context=candidate.panel_context,
            status="ERROR",
            website_score=candidate.score,
            checked_at=checked_at,
            notes="Verified listing could not be opened",
        )

    content = primary_content_locator(page)
    text = safe_inner_text(content, timeout=15_000)
    website_domain = canonical_domain(page.url) or candidate.domain
    emails = merge_contact_values(
        collect_emails(text, content, "verified_listing", website_domain)
    )
    phones = merge_contact_values(collect_phones(text, content, "verified_listing"))
    identity_ok, website_cuis, cui_match_status, identity_notes = identity_decision(
        expected_company,
        expected_cui,
        expected_county,
        expected_address,
        text,
    )

    if not identity_ok and cui_match_status == "MISMATCH":
        status = "REVIEW_CUI_MISMATCH"
        notes = identity_notes
    elif not identity_ok:
        status = "REVIEW_IDENTITY_MISMATCH"
        notes = identity_notes
    elif emails or phones:
        status = "FOUND_VERIFIED_LISTING"
        notes = f"Contacts extracted from verified listing; {identity_notes}"
    else:
        status = "WEBSITE_NO_CONTACT"
        notes = "No contact details found on the exact verified listing"

    keep_contacts = status == "FOUND_VERIFIED_LISTING"
    final_url = normalize_url(page.url) or candidate.url
    return ContactResult(
        website=final_url,
        google_source=candidate.source,
        google_title=candidate.title,
        google_query=candidate.query,
        panel_context=candidate.panel_context,
        website_cuis=website_cuis,
        cui_match_status=cui_match_status,
        contact_page_url=final_url,
        email=emails[0].value if emails and keep_contacts else "",
        phone=phones[0].value if phones and keep_contacts else "",
        all_emails=[item.value for item in emails] if keep_contacts else [],
        all_phones=[item.value for item in phones] if keep_contacts else [],
        found_in="verified_listing" if keep_contacts else "",
        status=status,
        website_score=candidate.score,
        checked_at=checked_at,
        notes=notes,
    )


def merge_contact_values(values: list[ContactValue]) -> list[ContactValue]:
    best: dict[str, ContactValue] = {}
    for item in values:
        existing = best.get(item.value)
        if existing is None or item.score > existing.score:
            best[item.value] = item
    return sorted(best.values(), key=lambda item: (-item.score, item.value))


def inspect_selected_website(
    page: Page,
    candidate: GoogleCandidate,
    expected_company: str,
    expected_cui: str,
    expected_county: str,
    expected_address: str,
) -> ContactResult:
    if is_verified_listing(candidate):
        return inspect_verified_listing(
            page,
            candidate,
            expected_company,
            expected_cui,
            expected_county,
            expected_address,
        )

    checked_at = datetime.now().isoformat(timespec="seconds")
    homepage = base_url(candidate.url)
    website_domain = candidate.domain

    print(f"  Selected {candidate.source} website: {homepage}")
    print("  Opening homepage/footer...")

    if not open_page(page, homepage):
        # Some sites reject the canonical root but allow the exact Google URL.
        if not open_page(page, candidate.url):
            return ContactResult(
                website=homepage,
                google_source=candidate.source,
                google_rank=candidate.rank,
                google_title=candidate.title,
                google_query=candidate.query,
                panel_context=candidate.panel_context,
                status="ERROR",
                website_score=candidate.score,
                duplicate_cuis=candidate.duplicate_cuis,
                checked_at=checked_at,
                notes="Selected website could not be opened",
            )

    homepage = base_url(page.url)
    website_domain = canonical_domain(page.url) or website_domain
    homepage_identity_text = safe_inner_text(page.locator("body"), timeout=15_000)
    footer_emails, footer_phones, _footer_cuis = footer_data(page, website_domain)
    contact_url = discover_contact_url(
        page=page,
        website_domain=website_domain,
        google_result_url=candidate.url,
    )

    contact_emails: list[ContactValue] = []
    contact_phones: list[ContactValue] = []
    _contact_cuis: list[str] = []
    final_contact_url = ""
    contact_identity_text = ""

    if contact_url:
        print(f"  Opening contact page: {contact_url}")
        (
            contact_emails,
            contact_phones,
            _contact_cuis,
            final_contact_url,
            contact_identity_text,
        ) = contact_page_data(page, contact_url, website_domain)
    else:
        contact_url = urljoin(homepage, "/contact")
        print(f"  No contact link found; trying standard path: {contact_url}")
        (
            contact_emails,
            contact_phones,
            _contact_cuis,
            final_contact_url,
            contact_identity_text,
        ) = contact_page_data(page, contact_url, website_domain)

    emails = merge_contact_values(contact_emails + footer_emails)
    phones = merge_contact_values(contact_phones + footer_phones)
    identity_text = f"{homepage_identity_text}\n{contact_identity_text}"
    identity_ok, website_cuis, cui_match_status, identity_notes = identity_decision(
        expected_company,
        expected_cui,
        expected_county,
        expected_address,
        identity_text,
    )
    if (
        not identity_ok
        and cui_match_status == "NOT_FOUND"
        and candidate.cui_found
        and candidate.company_score >= 65
    ):
        identity_ok = True
        cui_match_status = "GOOGLE_MATCH"
        identity_notes = (
            "Exact CUI and company name matched the Google result for this domain; "
            "the website publishes a different current address"
        )
    sources: list[str] = []
    if contact_emails or contact_phones:
        sources.append("contact_page")
    if footer_emails or footer_phones:
        sources.append("footer")

    if not identity_ok and cui_match_status == "MISMATCH":
        status = "REVIEW_CUI_MISMATCH"
    elif not identity_ok:
        status = "REVIEW_IDENTITY_MISMATCH"
    elif contact_emails or contact_phones:
        status = "FOUND_CONTACT_PAGE"
    elif footer_emails or footer_phones:
        status = "FOUND_FOOTER"
    else:
        status = "WEBSITE_NO_CONTACT"

    return ContactResult(
        website=homepage,
        google_source=candidate.source,
        google_rank=candidate.rank,
        google_title=candidate.title,
        google_query=candidate.query,
        panel_context=candidate.panel_context,
        website_cuis=website_cuis,
        cui_match_status=cui_match_status,
        contact_page_url=final_contact_url,
        email=emails[0].value if emails and identity_ok else "",
        phone=phones[0].value if phones and identity_ok else "",
        all_emails=(
            [item.value for item in emails] if identity_ok else []
        ),
        all_phones=(
            [item.value for item in phones] if identity_ok else []
        ),
        found_in="; ".join(sources) if identity_ok else "",
        status=status,
        website_score=candidate.score,
        duplicate_cuis=candidate.duplicate_cuis,
        checked_at=checked_at,
        notes=(
            f"Contacts rejected: {identity_notes}"
            if not identity_ok
            else "Website identity verified; "
            f"{identity_notes}; homepage footer and one contact page inspected"
        ),
    )


def find_input_columns(ws, header_row: int) -> dict[str, int | None]:
    headers = {
        normalize_header(cell.value): cell.column
        for cell in ws[header_row]
        if cell.value is not None
    }

    def first_alias(aliases: tuple[str, ...]) -> int | None:
        for alias in aliases:
            column = headers.get(normalize_header(alias))
            if column is not None:
                return column
        return None

    columns = {
        "company": first_alias(COMPANY_ALIASES),
        "cui": first_alias(CUI_ALIASES),
        "county": first_alias(COUNTY_ALIASES),
        "address": first_alias(ADDRESS_ALIASES),
    }
    if columns["company"] is None:
        raise ValueError("Company-name column not found")
    if columns["cui"] is None:
        raise ValueError("CUI column not found")
    return columns


def ensure_output_columns(ws, header_row: int) -> dict[str, int]:
    existing = {
        normalize_header(cell.value): cell.column
        for cell in ws[header_row]
        if cell.value is not None
    }
    result: dict[str, int] = {}
    next_column = ws.max_column + 1

    for name in OUTPUT_COLUMNS:
        normalized = normalize_header(name)
        if normalized in existing:
            result[name] = existing[normalized]
            continue
        result[name] = next_column
        cell = ws.cell(header_row, next_column, name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="164194")
        next_column += 1
    return result


def cell_text(ws, row: int, column: int | None) -> str:
    if column is None:
        return ""
    return str(ws.cell(row, column).value or "").strip()


def write_result(
    ws,
    row: int,
    output_columns: dict[str, int],
    result: ContactResult,
) -> None:
    values = {
        "google_website": result.website,
        "google_source": result.google_source,
        "google_result_rank": result.google_rank,
        "google_result_title": result.google_title,
        "google_query_used": result.google_query,
        "google_panel_context": result.panel_context,
        "website_cui": "; ".join(result.website_cuis),
        "cui_match_status": result.cui_match_status,
        "contact_page_url": result.contact_page_url,
        "contact_email": result.email,
        "contact_phone": result.phone,
        "contact_all_emails": "; ".join(result.all_emails),
        "contact_all_phones": "; ".join(result.all_phones),
        "contact_found_in": result.found_in,
        "contact_status": result.status,
        "website_score": result.website_score,
        "duplicate_domain_cui": "; ".join(result.duplicate_cuis),
        "contact_checked_at": result.checked_at,
        "contact_notes": result.notes,
    }
    for name, value in values.items():
        ws.cell(row, output_columns[name], value)

    color = STATUS_FILL_COLORS.get(result.status)
    if color:
        ws.cell(row, output_columns["contact_status"]).fill = PatternFill(
            "solid", fgColor=color
        )


def build_name_frequency(ws, columns: dict[str, int | None], header_row: int) -> Counter[str]:
    counter: Counter[str] = Counter()
    company_column = int(columns["company"])
    cui_column = int(columns["cui"])
    seen_pairs: set[tuple[str, str]] = set()

    for row in range(header_row + 1, ws.max_row + 1):
        company = cell_text(ws, row, company_column)
        cui = normalize_cui(ws.cell(row, cui_column).value)
        normalized = normalize_text(company)
        if not normalized or not cui or (normalized, cui) in seen_pairs:
            continue
        seen_pairs.add((normalized, cui))
        counter[normalized] += 1
    return counter


def build_assigned_domain_map(
    ws,
    input_columns: dict[str, int | None],
    output_columns: dict[str, int],
    header_row: int,
) -> dict[str, set[str]]:
    assigned: dict[str, set[str]] = defaultdict(set)
    cui_column = int(input_columns["cui"])

    for row in range(header_row + 1, ws.max_row + 1):
        website = cell_text(ws, row, output_columns["google_website"])
        status = cell_text(ws, row, output_columns["contact_status"])
        cui = normalize_cui(ws.cell(row, cui_column).value)
        domain = canonical_domain(website)
        if not domain or not cui or is_excluded_domain(domain):
            continue
        if status.startswith("FOUND_") or status == "WEBSITE_NO_CONTACT":
            assigned[domain].add(cui)
    return assigned


def output_path_for(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_google_contacts_v4.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Use Google's company panel first, otherwise select one organic "
            "website, then inspect only its contact page and homepage footer."
        )
    )
    parser.add_argument("input", nargs="?", type=Path)
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    parser.add_argument("--output", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--resume", action="store_true")
    parser.add_argument(
        "--retry-failed",
        action="store_true",
        help=(
            "With --resume, retry NO_WEBSITE, GOOGLE_BLOCKED, ERROR and "
            "review rows"
        ),
    )
    parser.add_argument("--only-cui")
    parser.add_argument("--max-results", type=int, default=10)
    parser.add_argument("--min-website-score", type=int, default=55)
    parser.add_argument("--google-delay", type=float, default=4.0)
    parser.add_argument("--company-delay", type=float, default=7.0)
    parser.add_argument("--manual-captcha", action="store_true", default=True)
    parser.add_argument(
        "--no-manual-captcha",
        dest="manual_captcha",
        action="store_false",
    )
    parser.add_argument("--save-every", type=int, default=1)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.input is None:
        print("Input Excel file is required.")
        return 2

    input_path = args.input.resolve()
    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 2

    output_path = (
        args.output.resolve() if args.output else output_path_for(input_path).resolve()
    )
    workbook_source = output_path if args.resume and output_path.exists() else input_path
    workbook = load_workbook(workbook_source)

    if args.sheet:
        if args.sheet not in workbook.sheetnames:
            print(f"Worksheet not found: {args.sheet}")
            return 2
        ws = workbook[args.sheet]
    else:
        ws = workbook.worksheets[0]

    try:
        input_columns = find_input_columns(ws, args.header_row)
    except ValueError as exc:
        print(f"ERROR: {exc}")
        return 2

    output_columns = ensure_output_columns(ws, args.header_row)
    name_frequency = build_name_frequency(ws, input_columns, args.header_row)
    assigned_domains = build_assigned_domain_map(
        ws, input_columns, output_columns, args.header_row
    )

    print(f"Google Contact Finder v{VERSION}")
    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print(f"Browser profile: {PROFILE_DIR}")
    print("Workflow: Google company panel first -> organic fallback -> one website")
    print("Website inspection: homepage footer + one contact page + CUI check")
    print(
        f"Limits: companies={args.limit or 'all'}, "
        f"results/query={args.max_results}, "
        f"minimum website score={args.min_website_score}"
    )
    if args.only_cui:
        print(f"Only CUI: {normalize_cui(args.only_cui)}")
    print()

    processed = 0
    status_counts: dict[str, int] = {}

    with sync_playwright() as playwright:
        context: BrowserContext = playwright.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            headless=False,
            locale="ro-RO",
            viewport={"width": 1400, "height": 900},
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.set_default_timeout(15_000)

        try:
            first_row = max(args.start_row, args.header_row + 1)
            for row in range(first_row, ws.max_row + 1):
                company = cell_text(ws, row, input_columns["company"])
                cui = normalize_cui(ws.cell(row, int(input_columns["cui"])).value)
                county = cell_text(ws, row, input_columns["county"])
                address = cell_text(ws, row, input_columns["address"])

                if not company or not cui:
                    continue
                if args.only_cui and cui != normalize_cui(args.only_cui):
                    continue

                existing_status = cell_text(
                    ws, row, output_columns["contact_status"]
                )
                if args.resume and existing_status:
                    retryable = existing_status in {
                        "NO_WEBSITE",
                        "WEBSITE_NO_CONTACT",
                        "GOOGLE_BLOCKED",
                        "ERROR",
                        "REVIEW_GOOGLE_CANDIDATE",
                        "REVIEW_DUPLICATE_DOMAIN",
                        "REVIEW_AMBIGUOUS_NAME",
                        "REVIEW_CUI_MISMATCH",
                        "REVIEW_PANEL_LISTING",
                        "REVIEW_IDENTITY_MISMATCH",
                    }
                    if not (args.retry_failed and retryable):
                        continue

                if args.limit > 0 and processed >= args.limit:
                    break

                print(
                    f"[{processed + 1}] Row {row}: {company} "
                    f"(CUI {cui}, county {county or '-'})"
                )
                checked_at = datetime.now().isoformat(timespec="seconds")

                try:
                    accepted, review, google_blocked, panel_contact = discover_website(
                        page=page,
                        company=company,
                        cui=cui,
                        county=county,
                        address=address,
                        max_results=args.max_results,
                        min_website_score=args.min_website_score,
                        manual_captcha=args.manual_captcha,
                        google_delay=args.google_delay,
                        name_frequency=name_frequency,
                        assigned_domains=assigned_domains,
                    )

                    if accepted is not None:
                        result = inspect_selected_website(
                            page,
                            accepted,
                            company,
                            cui,
                            county,
                            address,
                        )
                        if panel_contact is not None:
                            result = merge_google_panel_contact(result, panel_contact)
                    elif panel_contact is not None:
                        result = panel_contact
                    elif review is not None:
                        if review.duplicate_cuis:
                            status = "REVIEW_DUPLICATE_DOMAIN"
                        elif is_ambiguous_company(company, name_frequency):
                            status = "REVIEW_AMBIGUOUS_NAME"
                        else:
                            status = "REVIEW_GOOGLE_CANDIDATE"

                        result = ContactResult(
                            website=base_url(review.url),
                            google_source=review.source,
                            google_rank=review.rank,
                            google_title=review.title,
                            google_query=review.query,
                            panel_context=review.panel_context,
                            status=status,
                            website_score=review.score,
                            duplicate_cuis=review.duplicate_cuis,
                            checked_at=checked_at,
                            notes=review.rejection_reason,
                        )
                    elif google_blocked:
                        result = ContactResult(
                            status="GOOGLE_BLOCKED",
                            checked_at=checked_at,
                            notes="Google results unavailable after manual verification",
                        )
                    else:
                        result = ContactResult(
                            status="NO_WEBSITE",
                            checked_at=checked_at,
                            notes="No credible official website found",
                        )

                except KeyboardInterrupt:
                    raise
                except Exception as exc:
                    result = ContactResult(
                        status="ERROR",
                        checked_at=checked_at,
                        notes=f"{type(exc).__name__}: {exc}",
                    )

                write_result(ws, row, output_columns, result)
                processed += 1
                status_counts[result.status] = status_counts.get(result.status, 0) + 1

                if (
                    result.website
                    and not is_excluded_domain(canonical_domain(result.website))
                    and (result.status.startswith("FOUND_") or result.status == "WEBSITE_NO_CONTACT")
                ):
                    assigned_domains[canonical_domain(result.website)].add(cui)

                print(
                    f"  Result: {result.status} | "
                    f"source={result.google_source or '-'} | "
                    f"website={result.website or '-'} | "
                    f"email={result.email or '-'} | "
                    f"phone={result.phone or '-'} | "
                    f"score={result.website_score}"
                )

                if args.save_every > 0 and processed % args.save_every == 0:
                    workbook.save(output_path)
                    print(f"  Saved: {output_path.name}")

                if args.only_cui:
                    break
                time.sleep(max(0.0, args.company_delay))

        except KeyboardInterrupt:
            print()
            print("Interrupted. Saving current progress...")
        finally:
            workbook.save(output_path)
            context.close()

    print()
    print(f"Finished. Processed={processed}")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")
    print(f"Output: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
